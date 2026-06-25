from __future__ import annotations

import base64
import functools
import io
import os
import secrets
import uuid
from datetime import datetime
from pathlib import Path

from flask import (
    Flask,
    Response,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from werkzeug.utils import secure_filename

from utils.database import (
    fetch_analysis,
    get_dashboard_stats,
    get_report_rows,
    init_db,
    insert_analysis,
    insert_dataset_image,
    verify_user,
)
from utils.dataset_manager import (
    CATEGORY_GROUPS as DATASET_CATEGORY_GROUPS,
    clean_dataset,
    count_images,
    delete_image,
    ensure_structure,
    list_preview,
    save_unknown_image,
    save_uploaded_files,
    split_dataset,
)
from utils.detector import WasteDetector
from utils.economic_value import PRICE_PER_KG, calculate_economic_value
from utils.recommendation import build_recommendation
from utils.sort_dataset import sort_dataset
from utils.waste_catalog import get_item


BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "static" / "uploads"
CAPTURE_DIR = BASE_DIR / "static" / "captures"
DATASET_DIR = BASE_DIR / "dataset"
ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png"}

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("ECOLENS_SECRET_KEY") or secrets.token_hex(32)
app.config["MAX_CONTENT_LENGTH"] = 8 * 1024 * 1024

init_db(BASE_DIR / "database.db")
ensure_structure(DATASET_DIR)
detector = WasteDetector(model_path=BASE_DIR / "models" / "yolov8_waste.pt")


def login_required(view):
    """Proteksi halaman agar hanya user yang sudah login bisa mengakses dashboard."""

    @functools.wraps(view)
    def wrapped_view(**kwargs):
        if not session.get("is_logged_in"):
            return redirect(url_for("login"))
        return view(**kwargs)

    return wrapped_view


@app.template_filter("rupiah")
def rupiah(value: float | int | None) -> str:
    amount = int(value or 0)
    return "Rp{:,.0f}".format(amount).replace(",", ".")


@app.context_processor
def inject_dataset_status():
    """Tampilkan peringatan global jika dataset masih belum cukup."""
    try:
        return {"dataset_status": count_images(DATASET_DIR)}
    except Exception:
        return {"dataset_status": None}


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def public_path(path: Path) -> str:
    return path.relative_to(BASE_DIR).as_posix()


def analyze_image(image_path: Path, weight: float, source: str) -> dict:
    """Jalankan deteksi dan hitung nilai ekonomi tanpa menyimpan ke database."""
    detection = detector.analyze(image_path)
    if detection.get("needs_more_data"):
        try:
            save_unknown_image(image_path, DATASET_DIR)
        except Exception:
            pass
    recommendation = build_recommendation(detection["item_name"], detection["condition"], detection["item_key"])
    value_data = calculate_economic_value(detection["item_name"], detection["condition"], weight, detection["item_key"])

    record = {
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "image_path": public_path(image_path),
        "source": source,
        "item_key": detection["item_key"],
        "item_name": detection["item_name"],
        "waste_type": detection["item_name"],
        "category": detection["category"],
        "detected_label": detection["detected_label"],
        "confidence": detection["confidence"],
        "condition": detection["condition"],
        "potential": recommendation["potential"],
        "recycle_potential": recommendation["potential"],
        "weight": weight,
        "price_per_kg": value_data["price_per_kg"],
        "economic_value": value_data["economic_value"],
        "processing_action": recommendation["action"],
        "recommendation": recommendation["message"],
    }
    return record


def set_pending_analysis(record: dict) -> None:
    session["pending_analysis"] = record
    session.modified = True


def clear_pending_analysis() -> None:
    session.pop("pending_analysis", None)
    session.modified = True


@app.route("/")
def index():
    if session.get("is_logged_in"):
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if verify_user(username, password):
            session.clear()
            session["is_logged_in"] = True
            session["username"] = username
            flash("Selamat datang di EcoLens.", "success")
            return redirect(url_for("dashboard"))
        flash("Username atau password salah.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    stats = get_dashboard_stats()
    return render_template("dashboard.html", stats=stats)


@app.route("/analysis")
@login_required
def analysis():
    result = session.get("pending_analysis")
    saved = bool(session.pop("analysis_saved", False))
    return render_template(
        "analysis.html",
        result=result,
        saved=saved,
        item_groups=DATASET_CATEGORY_GROUPS,
    )


@app.route("/analysis/verify", methods=["POST"])
@login_required
def analysis_verify():
    pending = session.get("pending_analysis")
    if not pending:
        flash("Tidak ada hasil analisis yang perlu diverifikasi.", "error")
        return redirect(url_for("analysis"))
    pending["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    insert_analysis(pending)
    clear_pending_analysis()
    session["analysis_saved"] = True
    return redirect(url_for("analysis"))


@app.route("/analysis/correct", methods=["POST"])
@login_required
def analysis_correct():
    pending = session.get("pending_analysis")
    item_key = request.form.get("item_key", "").strip()
    item = get_item(item_key)
    if not pending or item is None:
        flash("Koreksi label tidak valid.", "error")
        return redirect(url_for("analysis"))

    weight = float(pending.get("weight") or 1)
    condition = pending.get("condition") or "Sedang"
    recommendation = build_recommendation(item.item_name, condition, item.item_key)
    value_data = calculate_economic_value(item.item_name, condition, weight, item.item_key)
    pending.update(
        {
            "item_key": item.item_key,
            "item_name": item.item_name,
            "waste_type": item.item_name,
            "category": item.category,
            "detected_label": f"Koreksi manual: {item.item_name}",
            "potential": recommendation["potential"],
            "recycle_potential": recommendation["potential"],
            "price_per_kg": value_data["price_per_kg"],
            "economic_value": value_data["economic_value"],
            "processing_action": recommendation["action"],
            "recommendation": recommendation["message"],
        }
    )
    set_pending_analysis(pending)
    flash("Label berhasil dikoreksi. Periksa kembali lalu klik Simpan & Verifikasi.", "success")
    return redirect(url_for("analysis"))


@app.route("/analysis/reset", methods=["POST"])
@login_required
def analysis_reset():
    clear_pending_analysis()
    flash("Hasil analisis lama sudah dibersihkan. Silakan pilih foto baru.", "success")
    return redirect(url_for("upload"))


@app.route("/upload", methods=["GET", "POST"])
@login_required
def upload():
    result = None
    if request.method == "POST":
        file = request.files.get("image")
        weight = float(request.form.get("weight") or 1)
        if not file or file.filename == "":
            flash("Pilih gambar sampah terlebih dahulu.", "error")
            return redirect(url_for("upload"))
        if not allowed_file(file.filename):
            flash("Format file harus JPG, JPEG, atau PNG.", "error")
            return redirect(url_for("upload"))

        filename = f"{datetime.now():%Y%m%d%H%M%S}_{uuid.uuid4().hex[:8]}_{secure_filename(file.filename)}"
        image_path = UPLOAD_DIR / filename
        file.save(image_path)
        result = analyze_image(image_path, weight, "Upload Foto")
        set_pending_analysis(result)
        flash("Foto berhasil dianalisis. Silakan verifikasi sebelum disimpan ke riwayat.", "success")
        return redirect(url_for("analysis"))

    return render_template("upload.html", result=result)


@app.route("/camera")
@login_required
def camera():
    return render_template("camera.html")


@app.route("/dataset", methods=["GET", "POST"])
@login_required
def dataset():
    if request.method == "POST":
        category = request.form.get("category", "")
        files = request.files.getlist("images")
        if not files or not any(file.filename for file in files):
            flash("Pilih minimal satu gambar dataset.", "error")
            return redirect(url_for("dataset"))
        try:
            summary = save_uploaded_files(files, category, DATASET_DIR)
            for saved_file in summary.get("files", []):
                try:
                    insert_dataset_image(category, saved_file.parent.parent.parent.name, saved_file.relative_to(BASE_DIR).as_posix())
                except Exception:
                    pass
            flash(
                f"Dataset tersimpan: {summary['saved']} gambar, duplikat: {summary['duplicates']}, tidak valid: {summary['invalid']}.",
                "success",
            )
        except ValueError as exc:
            flash(str(exc), "error")
        return redirect(url_for("dataset"))

    status = count_images(DATASET_DIR)
    previews = list_preview(DATASET_DIR)
    sort_report = session.pop("dataset_sort_report", None)
    return render_template("dataset.html", category_groups=DATASET_CATEGORY_GROUPS, status=status, previews=previews, sort_report=sort_report)


@app.route("/dataset/split", methods=["POST"])
@login_required
def dataset_split():
    split_dataset(DATASET_DIR)
    flash("Split dataset otomatis 70% train, 20% valid, 10% test selesai.", "success")
    return redirect(url_for("dataset"))


@app.route("/dataset/clean", methods=["POST"])
@login_required
def dataset_clean():
    result = clean_dataset(DATASET_DIR)
    flash(
        f"Clean dataset selesai. Gambar rusak: {result['removed_invalid']}, duplikat: {result['removed_duplicates']}.",
        "success",
    )
    return redirect(url_for("dataset"))


@app.route("/dataset/sort", methods=["POST"])
@login_required
def dataset_sort():
    report = sort_dataset()
    session["dataset_sort_report"] = report
    flash(
        f"Auto pilah selesai: {report['specific']} gambar masuk jenis spesifik, {report['unknown']} masuk unknown.",
        "success",
    )
    return redirect(url_for("dataset"))


@app.route("/dataset/delete/<split>/<category>/<item_key>/<filename>", methods=["POST"])
@login_required
def dataset_delete(split: str, category: str, item_key: str, filename: str):
    if delete_image(split, category, filename, DATASET_DIR, item_key=item_key):
        flash("Gambar dataset berhasil dihapus.", "success")
    else:
        flash("Gambar dataset tidak ditemukan.", "error")
    return redirect(url_for("dataset"))


@app.route("/dataset/delete/unknown/<filename>", methods=["POST"])
@login_required
def dataset_delete_unknown(filename: str):
    if delete_image("unknown", "unknown", filename, DATASET_DIR):
        flash("Gambar unknown berhasil dihapus.", "success")
    else:
        flash("Gambar unknown tidak ditemukan.", "error")
    return redirect(url_for("dataset"))


@app.route("/dataset/file/<split>/<category>/<item_key>/<filename>")
@login_required
def dataset_file(split: str, category: str, item_key: str, filename: str):
    directory = DATASET_DIR / split / category / item_key
    return send_from_directory(directory, filename)


@app.route("/dataset/file/unknown/<filename>")
@login_required
def dataset_unknown_file(filename: str):
    return send_from_directory(DATASET_DIR / "unknown", filename)


@app.route("/camera/analyze", methods=["POST"])
@login_required
def analyze_camera_image():
    payload = request.get_json(silent=True) or {}
    image_data = payload.get("image_data", "")
    weight = float(payload.get("weight") or 1)
    if "," not in image_data:
        return jsonify({"ok": False, "message": "Data gambar kamera tidak valid."}), 400

    header, encoded = image_data.split(",", 1)
    image_bytes = base64.b64decode(encoded)
    filename = f"camera_{datetime.now():%Y%m%d%H%M%S}_{uuid.uuid4().hex[:8]}.png"
    image_path = CAPTURE_DIR / filename
    image_path.write_bytes(image_bytes)
    result = analyze_image(image_path, weight, "Webcam Browser")
    set_pending_analysis(result)
    return jsonify({"ok": True, "redirect": url_for("analysis"), "result": result})


@app.route("/camera/opencv-capture", methods=["POST"])
@login_required
def opencv_capture():
    """Capture langsung dari kamera laptop lokal menggunakan OpenCV."""
    try:
        import cv2
    except Exception:
        flash("OpenCV belum tersedia. Jalankan pip install -r requirements.txt.", "error")
        return redirect(url_for("camera"))

    weight = float(request.form.get("weight") or 1)
    camera = cv2.VideoCapture(0)
    ok, frame = camera.read()
    camera.release()
    if not ok:
        flash("Kamera laptop tidak dapat diakses oleh OpenCV.", "error")
        return redirect(url_for("camera"))

    filename = f"opencv_{datetime.now():%Y%m%d%H%M%S}_{uuid.uuid4().hex[:8]}.jpg"
    image_path = CAPTURE_DIR / filename
    cv2.imwrite(str(image_path), frame)
    result = analyze_image(image_path, weight, "OpenCV Camera")
    set_pending_analysis(result)
    flash("Foto dari OpenCV berhasil dianalisis. Silakan verifikasi sebelum disimpan ke riwayat.", "success")
    return redirect(url_for("analysis"))


@app.route("/history")
@login_required
def history():
    rows = fetch_analysis(limit=100)
    return render_template("history.html", rows=rows)


@app.route("/report")
@login_required
def report():
    start_date = request.args.get("start_date") or ""
    end_date = request.args.get("end_date") or ""
    rows = get_report_rows(start_date, end_date)
    total_value = sum(float(row["economic_value"] or 0) for row in rows)
    categories = {}
    for row in rows:
        categories[row["category"]] = categories.get(row["category"], 0) + 1
    return render_template(
        "report.html",
        rows=rows,
        total_value=total_value,
        categories=categories,
        start_date=start_date,
        end_date=end_date,
    )


@app.route("/report/export/xlsx")
@login_required
def export_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    rows = get_report_rows(request.args.get("start_date") or "", request.args.get("end_date") or "")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Laporan EcoLens"
    headers = [
        "Tanggal",
        "Jenis Spesifik",
        "Kategori",
        "Kondisi",
        "Potensi",
        "Berat (kg)",
        "Nilai Ekonomi",
        "Rekomendasi",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="147A45")
    for row in rows:
        sheet.append(
            [
                row["created_at"],
                row.get("item_name") or row.get("waste_type"),
                row["category"],
                row["condition"],
                row["potential"],
                row["weight"],
                row["economic_value"],
                row["processing_action"],
            ]
        )
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 22
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="laporan-ecolens.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/report/export/pdf")
@login_required
def export_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    rows = get_report_rows(request.args.get("start_date") or "", request.args.get("end_date") or "")
    output = io.BytesIO()
    document = SimpleDocTemplate(output, pagesize=landscape(A4), title="Laporan EcoLens")
    styles = getSampleStyleSheet()
    elements = [Paragraph("Laporan Analisis Sampah EcoLens", styles["Title"]), Spacer(1, 12)]
    data = [["Tanggal", "Jenis Spesifik", "Kategori", "Kondisi", "Potensi", "Berat", "Nilai", "Aksi"]]
    for row in rows:
        data.append(
            [
                row["created_at"],
                row.get("item_name") or row.get("waste_type"),
                row["category"],
                row["condition"],
                row["potential"],
                f"{row['weight']} kg",
                rupiah(row["economic_value"]),
                row["processing_action"],
            ]
        )
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#147A45")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9E5DE")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    elements.append(table)
    document.build(elements)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="laporan-ecolens.pdf", mimetype="application/pdf")


@app.route("/about")
@login_required
def about():
    return render_template("about.html", prices=PRICE_PER_KG)


if __name__ == "__main__":
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    CAPTURE_DIR.mkdir(parents=True, exist_ok=True)
    debug_enabled = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(debug=debug_enabled, host="127.0.0.1", port=5000)
